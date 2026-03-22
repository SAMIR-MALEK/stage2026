"""لوحة لجنة الانتقاء — مبنية من الصفر بأعمدة الشيت الجديدة"""
import streamlit as st
import pandas as pd
import json
from pathlib import Path
from datetime import datetime

RANK_SCORE_RANGES = {
    "الإقامة العلمية قصيرة المدى":                  (3.0,  9.0),
    "تربص تحسين المستوى":                            (3.0,  9.0),
    "تداريب تحسين المستوى - طلبة الدكتوراه":        (3.0,  9.0),
    "تربصات تحسين المستوى - إداريون وتقنيون":       (8.0, 12.0),
    "أساتذة محاضرون":                                (3.0,  9.0),
    "أساتذة مساعدون":                                (3.0,  9.0),
    "طلبة دكتوراه":                                  (3.0,  9.0),
    "إداريون وتقنيون":                               (8.0, 12.0),
}
STATUS_OPTIONS = ["قيد المراجعة", "مقبول", "مرفوض", "قائمة انتظار"]

LABEL_MAP = {
    "tr_rank":"وثيقة الرتبة العلمية","sc_rank":"وثيقة الرتبة العلمية",
    "rs_rank":"وثيقة الرتبة","rank_doc":"وثيقة الرتبة الوظيفية",
    "tr_reg_doc":"وثيقة التسجيل","tr_award_doc":"وثيقة الجائزة",
    "tr_sup_cert_0":"محضر دكتوراه 1","tr_sup_cert_1":"محضر دكتوراه 2",
    "tr_int_cert_0":"شهادة مداخلة 1","tr_int_cert_1":"شهادة مداخلة 2",
    "tr_pat_cert_0":"براءة اختراع 1","tr_proj_cert_0":"وثيقة مشروع 1",
    "tr_master_doc_1":"محاضر مناقشة ماستر","tr_lic_doc_1":"وثيقة إشراف ليسانس",
    "tr_shared_doc_1":"وثيقة جذع مشترك","high_doc":"وثيقة المنصب العالي",
    "adm_f1_istimara":"استمارة الترشح","adm_f1_mashrou3":"مشروع العمل","adm_f1_ta3ahod":"التعهد",
    "adm_f2_istimara":"استمارة الترشح","adm_f2_mashrou3":"مشروع العمل","adm_f2_ta3ahod":"التعهد",
    "adm_f3_istimara":"استمارة الترشح","adm_f3_mashrou3":"مشروع العمل",
    "adm_f3_ta3ahod":"التعهد","adm_f3_tasrih":"التصريح الشرفي",
    "adm_f4_istimara":"استمارة الترشح","adm_f4_mashrou3":"مشروع العمل","adm_f4_ta3ahod":"التعهد",
    "body_cert_0":"شهادة هيئة مرافقة 1","iproj_cert_0":"شهادة مشروع دولي 1",
}


def _logout():
    for k in list(st.session_state.keys()): del st.session_state[k]
    st.rerun()


def _safe_float(v):
    try:
        return float(str(v).replace(",","").replace(" ","")) if v not in ("","None",None) else 0.0
    except:
        return 0.0


def _load_all() -> list[dict]:
    records = []
    try:
        from utils.sheets import _get_client, SHEET_NAME
        cl = _get_client()
        if cl:
            recs = cl.open(SHEET_NAME).sheet1.get_all_records()
            if recs: return recs
    except Exception:
        pass
    sub = Path("data/submissions")
    if sub.exists():
        for f in sorted(sub.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                with open(f, encoding="utf-8") as fp:
                    d = json.load(fp); d["_file"] = str(f); records.append(d)
            except Exception:
                pass
    return records


def _normalize(raw: list[dict]) -> pd.DataFrame:
    rows = []
    for r in raw:
        rows.append({
            "اسم_المستخدم":   str(r.get("اسم_المستخدم", r.get("username",""))).strip(),
            "الاسم_الكامل":   str(r.get("الاسم_الكامل", r.get("name",""))).strip(),
            "السلك":          str(r.get("السلك",  r.get("silk",""))).strip(),
            "الرتبة":         str(r.get("الرتبة", r.get("rank",""))).strip(),
            "الصيغة":         str(r.get("الصيغة", r.get("scale",""))).strip(),
            "النقاط_الجزئية": _safe_float(r.get("النقاط_الجزئية", r.get("total_score",0))),
            "نقاط_الرتبة":   _safe_float(r.get("نقاط_الرتبة",  r.get("rank_pts",0))),
            "النقاط_الكلية": _safe_float(r.get("النقاط_الكلية", r.get("total_score",0))),
            "الحالة":         str(r.get("الحالة",  r.get("status","قيد المراجعة"))).strip(),
            "التاريخ":        str(r.get("التاريخ","")).strip(),
            "التفصيل":        str(r.get("تفصيل_النقاط", r.get("breakdown","{}")) or "{}"),
            "روابط_Drive":    str(r.get("روابط_الوثائق", r.get("drive_links","{}")) or "{}"),
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _save(username: str, rank_pts: float, status: str) -> bool:
    try:
        from utils.sheets import _get_client, SHEET_NAME
        cl = _get_client()
        if cl:
            ws   = cl.open(SHEET_NAME).sheet1
            recs = ws.get_all_records()
            hdrs = list(recs[0].keys()) if recs else []
            for idx, row in enumerate(recs, 2):
                if str(row.get("اسم_المستخدم","")).strip() == username:
                    partial = _safe_float(row.get("النقاط_الجزئية", 0))
                    total   = partial + rank_pts
                    if "النقاط_الرتبة" in hdrs:
                        ws.update_cell(idx, hdrs.index("النقاط_الرتبة")+1, rank_pts)
                    if "النقاط_الكلية" in hdrs:
                        ws.update_cell(idx, hdrs.index("النقاط_الكلية")+1, total)
                    if "الحالة" in hdrs:
                        ws.update_cell(idx, hdrs.index("الحالة")+1, status)
                    return True
    except Exception:
        pass
    sub = Path("data/submissions")
    if sub.exists():
        for f in sub.glob(f"{username}_*.json"):
            try:
                with open(f, encoding="utf-8") as fp: d = json.load(fp)
                d["rank_pts"]    = rank_pts
                d["total_score"] = _safe_float(d.get("total_score",0)) + rank_pts
                d["status"]      = status
                with open(f, "w", encoding="utf-8") as fp:
                    json.dump(d, fp, ensure_ascii=False, indent=2)
                return True
            except Exception: pass
    return False


def show_committee():
    st.markdown(f"""
    <div class="gov-header">
      <div style="font-size:.74rem;opacity:.6;margin-bottom:.2rem;">
        الجمهورية الجزائرية الديمقراطية الشعبية — وزارة التعليم العالي والبحث العلمي
      </div>
      <h1>👥 لوحة لجنة الانتقاء</h1>
      <p>كلية الحقوق والعلوم السياسية — جامعة محمد البشير الإبراهيمي برج بوعريريج</p>
      <span class="badge b-blue">{st.session_state.user_name}</span>
    </div>
    """, unsafe_allow_html=True)

    _, c2 = st.columns([5,1])
    with c2:
        if st.button("🚪 خروج", use_container_width=True): _logout()

    with st.spinner("جارٍ تحميل الملفات..."):
        raw = _load_all()
        df  = _normalize(raw)

    if df.empty:
        st.markdown('<div class="alert al-in">📭 لا توجد ملفات مقدَّمة بعد.</div>', unsafe_allow_html=True)
        return

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 قائمة المترشحين", "✏️ نقاط الرتبة", "🔧 تعديل النقاط", "✅ النتائج", "📥 تصدير"
    ])

    # ════ تبويب 1 ════
    with tab1:
        c1,c2,c3,c4 = st.columns(4)
        for col, (val, color, label) in zip([c1,c2,c3,c4], [
            (len(df), "#1a3a5c", "إجمالي الملفات"),
            (len(df[df["الحالة"]=="مقبول"]), "#27ae60", "مقبول"),
            (len(df[df["الحالة"]=="قيد المراجعة"]), "#c8973a", "قيد المراجعة"),
            (len(df[df["نقاط_الرتبة"]==0]), "#e74c3c", "بدون نقاط رتبة"),
        ]):
            with col:
                st.markdown(f'<div class="card"><div style="font-size:1.8rem;font-weight:800;color:{color};">{val}</div><div style="font-size:.82rem;color:#6b7f96;">{label}</div></div>', unsafe_allow_html=True)

        # فلتر
        silks    = ["الكل"] + sorted(df["السلك"].dropna().unique().tolist())
        sel_silk = st.selectbox("تصفية حسب السلك:", silks)
        filtered = df.copy() if sel_silk=="الكل" else df[df["السلك"]==sel_silk].copy()
        filtered = filtered.sort_values("النقاط_الكلية", ascending=False).reset_index(drop=True)
        filtered.index += 1

        def _color(val):
            m = {"مقبول":"background-color:#e8f8f0;color:#1a7a4a",
                 "مرفوض":"background-color:#fdecea;color:#a93226",
                 "قائمة انتظار":"background-color:#e6f1fb;color:#185fa5",
                 "قيد المراجعة":"background-color:#fef9ec;color:#7d6010"}
            return m.get(val,"")

        show_cols = ["الاسم_الكامل","السلك","الصيغة","النقاط_الجزئية","نقاط_الرتبة","النقاط_الكلية","الحالة"]
        st.dataframe(
            filtered[show_cols].style.applymap(_color, subset=["الحالة"]),
            use_container_width=True, height=380
        )

        st.markdown("---")
        names    = ["—"] + filtered["الاسم_الكامل"].tolist()
        sel_name = st.selectbox("عرض تفاصيل ملف:", names)
        if sel_name != "—":
            matches = filtered[filtered["الاسم_الكامل"]==sel_name]
            if not matches.empty:
                _detail(matches.iloc[0])

    # ════ تبويب 2 ════
    with tab2:
        st.markdown('<div class="alert al-wn">راجع وثيقة آخر ترقية ثم أدخل نقاط الرتبة.</div>', unsafe_allow_html=True)
        needs = df[df["نقاط_الرتبة"]==0].copy()
        if needs.empty:
            st.markdown('<div class="alert al-ok">✅ تمت إضافة نقاط الرتبة لجميع المترشحين.</div>', unsafe_allow_html=True)
        else:
            for _, row in needs.iterrows():
                silk  = row.get("السلك","")
                scale = row.get("الصيغة", silk)
                rng   = RANK_SCORE_RANGES.get(scale, RANK_SCORE_RANGES.get(silk, (0.0, 12.0)))
                has   = float(rng[1]) > 0

                st.markdown('<div class="item-block">', unsafe_allow_html=True)
                c1,c2,c3,c4 = st.columns([3,2,1.5,1.5])
                with c1:
                    st.markdown(f"**{row['الاسم_الكامل']}**")
                    st.markdown(f'<span style="font-size:.8rem;color:#6b7f96;">{silk} — {scale}</span>', unsafe_allow_html=True)
                    # رابط وثيقة الرتبة
                    try:
                        links = json.loads(row.get("روابط_Drive","{}") or "{}")
                        rank_link = links.get("وثيقة_الرتبة_العلمية",
                                    links.get("وثيقة_الرتبة_الوظيفية",
                                    links.get("tr_rank", links.get("rank_doc",""))))
                        if rank_link and str(rank_link).startswith("http"):
                            st.markdown(f'<a href="{rank_link}" target="_blank" style="font-size:.8rem;color:#185fa5;text-decoration:underline;">🔗 وثيقة الرتبة</a>', unsafe_allow_html=True)
                    except Exception:
                        pass
                with c2:
                    st.markdown(f"نقاط جزئية: **{row['النقاط_الجزئية']:.1f}**")
                with c3:
                    rank_val = st.number_input(
                        f"نقاط الرتبة ({float(rng[0]):.0f}–{float(rng[1]):.0f})",
                        min_value=float(rng[0]),
                        max_value=float(rng[1]) if float(rng[1]) > 0 else 0.1,
                        value=float(rng[0]),
                        step=0.5,
                        key=f"ri_{row['اسم_المستخدم']}",
                        disabled=not has
                    ) if has else 0.0
                with c4:
                    total = row["النقاط_الجزئية"] + rank_val
                    st.markdown(f'<div style="font-weight:800;color:#1a3a5c;font-size:1.1rem;text-align:center;">= {total:.1f} ن</div>', unsafe_allow_html=True)
                    if st.button("💾 حفظ", key=f"sv_{row['اسم_المستخدم']}", use_container_width=True):
                        if _save(row["اسم_المستخدم"], rank_val, row["الحالة"]):
                            st.success("✅"); st.rerun()
                        else:
                            st.error("❌ فشل الحفظ")
                st.markdown('</div>', unsafe_allow_html=True)

    # ════ تبويب 5 ════
    with tab5:
        st.markdown('<div class="alert al-in">حدّد نتيجة كل مترشح بعد مراجعة ملفه.</div>', unsafe_allow_html=True)
        df_s = df.sort_values("النقاط_الكلية", ascending=False).reset_index(drop=True)
        for idx, row in df_s.iterrows():
            c1,c2,c3,c4 = st.columns([3,1.5,2,1.5])
            with c1:
                st.markdown(
                    f'<div style="font-weight:600;">#{idx+1} {row["الاسم_الكامل"]}'
                    f'<br><span style="font-size:.78rem;color:#6b7f96;">{row.get("السلك","")} — {row.get("الصيغة","")}</span></div>',
                    unsafe_allow_html=True)
            with c2:
                st.markdown(
                    f'<div style="text-align:center;"><div style="font-size:1.4rem;font-weight:800;color:#1a3a5c;">{row["النقاط_الكلية"]:.1f}</div>'
                    f'<div style="font-size:.72rem;color:#6b7f96;">نقطة</div></div>',
                    unsafe_allow_html=True)
            with c3:
                cur_idx = STATUS_OPTIONS.index(row["الحالة"]) if row["الحالة"] in STATUS_OPTIONS else 0
                new_st  = st.selectbox("النتيجة", STATUS_OPTIONS, index=cur_idx,
                                       key=f"st_{row['اسم_المستخدم']}_{idx}",
                                       label_visibility="collapsed")
            with c4:
                if st.button("💾", key=f"svst_{row['اسم_المستخدم']}_{idx}", use_container_width=True):
                    if _save(row["اسم_المستخدم"], row["نقاط_الرتبة"], new_st):
                        st.success("✅"); st.rerun()
                    else:
                        st.error("❌")
            st.markdown('<hr style="margin:.3rem 0;border-color:#eee;">', unsafe_allow_html=True)

        if st.button("💾 حفظ جميع النتائج دفعة واحدة", use_container_width=True):
            n = 0
            for idx, row in df_s.iterrows():
                st_val = st.session_state.get(f"st_{row['اسم_المستخدم']}_{idx}", row["الحالة"])
                if _save(row["اسم_المستخدم"], row["نقاط_الرتبة"], st_val): n += 1
            st.success(f"✅ تم حفظ {n} نتيجة"); st.rerun()

    # ════ تبويب 5 ════
    with tab5:
        ex = df.sort_values("النقاط_الكلية", ascending=False).reset_index(drop=True)
        ex.index += 1; ex.index.name = "الترتيب"
        show = ["الاسم_الكامل","السلك","الصيغة","النقاط_الجزئية","نقاط_الرتبة","النقاط_الكلية","الحالة"]
        st.dataframe(ex[show], use_container_width=True)

        c1,c2 = st.columns(2)
        with c1:
            st.download_button(
                "📥 تحميل CSV",
                data=ex[show].to_csv(encoding="utf-8-sig", index=True).encode("utf-8-sig"),
                file_name=f"المترشحون_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv", use_container_width=True
            )
        with c2:
            try:
                import io, openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "قائمة المترشحين"
                hdrs = ["الترتيب"] + show
                for ci, h in enumerate(hdrs, 1):
                    c = ws.cell(1, ci, h)
                    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                    c.fill = PatternFill("solid", fgColor="1A3A5C")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    ws.column_dimensions[get_column_letter(ci)].width = 20
                STATUS_FILL = {"مقبول":"E8F8F0","مرفوض":"FDECEA","قائمة انتظار":"E6F1FB","قيد المراجعة":"FEF9EC"}
                for ri, (i, row) in enumerate(ex.iterrows(), 2):
                    vals = [i] + [row.get(c,"") for c in show]
                    fc   = PatternFill("solid", fgColor=STATUS_FILL.get(str(row.get("الحالة","")),"FFFFFF"))
                    for ci, v in enumerate(vals, 1):
                        cell = ws.cell(ri, ci, v)
                        cell.fill = fc
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                buf = io.BytesIO(); wb.save(buf)
                st.download_button(
                    "📥 تحميل Excel", data=buf.getvalue(),
                    file_name=f"المترشحون_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"خطأ Excel: {e}")


def _detail(row):
    """تفاصيل ملف مترشح"""
    username = str(row.get("اسم_المستخدم",""))
    st.markdown(f"""
    <div class="card">
      <div class="card-title">📋 ملف: {row['الاسم_الكامل']}</div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:.5rem;margin-bottom:.8rem;">
        <div><span style="color:#6b7f96;font-size:.8rem;">السلك</span><br><strong>{row.get('السلك','')}</strong></div>
        <div><span style="color:#6b7f96;font-size:.8rem;">الصيغة</span><br><strong>{row.get('الصيغة','')}</strong></div>
        <div><span style="color:#6b7f96;font-size:.8rem;">تاريخ التقديم</span><br><strong>{row.get('التاريخ','')}</strong></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # تفصيل النقاط
    try:
        bd = row.get("التفصيل","{}") or "{}"
        breakdown = json.loads(bd) if isinstance(bd,str) else bd
        if breakdown:
            st.markdown('<div class="card"><div class="card-title">📊 تفصيل النقاط</div>', unsafe_allow_html=True)
            for k, v in breakdown.items():
                if v is None:
                    st.markdown(f'<div class="score-row"><span>{k}</span><span style="color:#c8973a;">⏳ اللجنة</span></div>', unsafe_allow_html=True)
                else:
                    try:
                        fv  = float(v)
                        col = "#e74c3c" if fv < 0 else "#1a3a5c"
                        st.markdown(f'<div class="score-row"><span>{k}</span><span style="font-weight:700;color:{col};">{fv:+.1f} ن</span></div>', unsafe_allow_html=True)
                    except Exception:
                        pass
            c1,c2,c3 = st.columns(3)
            for col,(label,val,color) in zip([c1,c2,c3],[
                ("نقاط جزئية", row["النقاط_الجزئية"], "#c8973a"),
                ("نقاط الرتبة", row["نقاط_الرتبة"],   "#185fa5"),
                ("المجموع",     row["النقاط_الكلية"],  "#27ae60")
            ]):
                with col:
                    st.markdown(f'<div style="text-align:center;padding:.7rem;background:#f8f9fa;border-radius:8px;"><div style="font-size:.78rem;color:#6b7f96;">{label}</div><div style="font-size:1.4rem;font-weight:800;color:{color};">{float(val):.1f}</div></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
    except Exception:
        pass

    # وثائق Drive
    try:
        links_raw = row.get("روابط_Drive","{}") or "{}"
        links = json.loads(links_raw) if isinstance(links_raw,str) else links_raw
        if links and any(v for v in links.values()):
            st.markdown('<div class="card"><div class="card-title">📎 وثائق المترشح</div>', unsafe_allow_html=True)
            for doc_name, link in links.items():
                display = LABEL_MAP.get(doc_name, doc_name.replace("_"," "))
                if link and str(link).startswith("http"):
                    c1,c2 = st.columns([4,1])
                    with c1:
                        st.markdown(f"📄 **{display}**")
                    with c2:
                        st.markdown(f'<a href="{link}" target="_blank" style="display:inline-block;padding:4px 12px;background:#1a3a5c;color:white;border-radius:6px;font-size:.8rem;text-decoration:none;">🔗 فتح</a>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
    except Exception:
        pass
